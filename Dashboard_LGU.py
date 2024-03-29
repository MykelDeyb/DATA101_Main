import dash
import plotly.express as px
import dash_bootstrap_components as dbc
from dash import dcc, html, Input, Output
from openpyxl import load_workbook
from pathlib import Path

dataset_folder = Path('Datasets/')
workbook = load_workbook(dataset_folder / 'LGU_Data/LGUs.xlsx')

pillar_data = {}

for sheet in workbook:
    provinces = []
    scores = []
    distances_km = []
    distances_mi = []
    categories = [] 
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if len(row) >= 13:  
            provinces.append(row[0])
            scores.append(row[1:11])
            distances_km.append(row[11])
            distances_mi.append(row[12])
            categories.append(row[13])  

    pillar_name = sheet.title
    pillar_data[pillar_name] = {
        'provinces': provinces,
        'scores': scores,
        'distances_km': distances_km,
        'distances_mi': distances_mi,
        'categories': categories 
    }

pillar_names = list(pillar_data.keys())

all_years = list(range(2014, 2024))

provinces = list(set(province for pillar in pillar_names for province in pillar_data[pillar]['provinces']))
pillar_descriptions = {
    'overall score': {
        'Description': 'Overall score is the sum of scores on five main pillars which pool data from several sub-indicators. The higher the score of a city or municipality, the more competitive it is.'
    },
    'economic dynamism': {
        'Description': 'Creates stable expansion of businesses and industries and higher employment. Matches output and productivity of the local economy with the local resources. Localities are centers of economic activities, and due to this, business expansion and job creation are easily observable in local settings.'
    },
    'government efficiency': {
        'Description': 'Refers to the quality and reliability of government services and government support for effective and sustainable productive expansion. This factor looks at government as an institution that is generally not corrupt; able to protect and enforce contracts; apply moderate and reasonable taxation and is able to regulate proactively.'
    },
    'infrastructure': {
        'Description': 'Pertains to the physical assets that connect, expand, and sustain a locality and its surroundings to enable provision of goods and services. It involves basic inputs of production such as energy, water; interconnection of production such as transportation, roads and communications; sustenance of production such as waste, disaster preparedness, environmental sustainability; and human capital formation infrastructure.'
    },
    'resiliency': {
        'Description': 'Applies to the capacity of a locality to build systems that can absorb change and disturbance and being able to adapt to such changes. It spans frameworks that bind LGUs and their constituents to prepare for possible shocks and stresses; budgeting for disaster risk reduction; hazard/risk identification mechanisms; resilience-related infrastructure; and resilience-related mechanisms.'
    },
    'innovation': {
        'Description': 'Refers to the ability of a locality to harness its creative potential to improve or sustain current levels of productivity. It hinges mainly on the development of creative capital which are human resources, research capabilities, and networking capacities.'
    }
}

external_stylesheets = ['styles.css']

app = dash.Dash(__name__, external_stylesheets=external_stylesheets)

app.layout = html.Div([
    # Header
    html.Div([
        html.H1("Dashboard", style={'display': 'inline-block', 'margin-bottom': '20px', 'text-align': 'center'}),
        html.Div([
            # Level Dropdown
            html.Label('Level'),
            dcc.Dropdown(
                id='level-dropdown',
                options=[
                    {'label': 'Province', 'value': 'Province'},
                    {'label': 'LGU', 'value': 'LGU'}
                ],
                value='LGU',
                style={'width': '100px', 'display': 'inline-block', 'margin-left': '20px', 'margin-right': '20px'}
            ),
            # Pillar Dropdown
            html.Label('Pillar'),
            dcc.Dropdown(
                id='pillar-dropdown',
                options=[{'label': pillar, 'value': pillar} for pillar in pillar_names],
                value=pillar_names[0],
                style={'width': '150px', 'display': 'inline-block', 'margin-right': '20px'}
            ),
            # Starting Year Dropdown
            html.Label('Starting Year'),
            dcc.Dropdown(
                id='start-year-dropdown',
                options=[{'label': str(year), 'value': year} for year in all_years],
                value=2014,
                style={'width': '80px', 'display': 'inline-block', 'margin-right': '20px'}
            ),
            # End Year Dropdown
            html.Label('End Year'),
            dcc.Dropdown(
                id='end-year-dropdown',
                options=[{'label': str(year), 'value': year} for year in all_years],
                value=2023,
                style={'width': '80px', 'display': 'inline-block'}
            ),
        ], style={'margin-bottom': '20px', 'text-align': 'center'}),
    ], style={'text-align': 'center'}),

    html.Div([
        html.Div([
            # Sidebar
            html.Div([
                # Search bar
                html.Label('Select Provinces'),
                dcc.Input(id='province-search', type='text', placeholder='Search Provinces...'),
                dcc.Checklist(
                    id='province-checkboxes',
                    options=[{'label': province, 'value': province} for province in provinces],
                    value=[],
                    style={'overflowY': 'scroll', 'height': '400px'}
                ),
                html.Button('Clear Selection', id='clear-selection-button', n_clicks=0)
            ], style={'width': '20%', 'float': 'right', 'margin-left': '20px'}),
        ]),
        # Table province data
        html.Div(id='table-container', style={'width': '75%', 'float': 'left'}),

        # Line chart
        dcc.Graph(id='line-chart', style={'width': '75%', 'float': 'left'}),
        # Composition of Overall Score Bar chart
        html.Div([
            html.H3('Composition of Overall Score', style={'text-align': 'center'}),
            html.Label('Pillar'),
            dcc.Dropdown(
                id='bar-pillar-dropdown',
                options=[{'label': pillar, 'value': pillar} for pillar in pillar_names],
                value=pillar_names[0],
                style={'width': '150px', 'margin-right': '20px'}
            ),
            html.Label('Year'),
            dcc.Dropdown(
                id='bar-year-dropdown',
                options=[{'label': str(year), 'value': year} for year in all_years],
                value=2014,
                style={'width': '80px'}
            ),
            dcc.Graph(id='bar-chart')
        ], style={'width': '75%', 'float': 'left', 'margin-top': '20px'}),

        # Pillar Information
        html.Div([
            html.H3('Pillar Information', style={'text-align': 'center'}),
            html.Div(id='pillar-info-container')
        ], style={'width': '75%', 'float': 'left', 'margin-top': '20px'})
    ])
])


@app.callback(
    Output('province-checkboxes', 'options'),
    [Input('province-search', 'value')]
)
def update_province_options(search_value):
    if search_value is None:
        sorted_provinces = sorted(provinces)
        return [{'label': province, 'value': province} for province in sorted_provinces]
    else:
        filtered_provinces = [province for province in provinces if search_value.lower() in province.lower()]
        sorted_filtered_provinces = sorted(filtered_provinces)
        return [{'label': province, 'value': province} for province in sorted_filtered_provinces]


@app.callback(
    Output('province-checkboxes', 'value'),
    [Input('clear-selection-button', 'n_clicks')]
)
def clear_selected_provinces(n_clicks):
    if n_clicks > 0:
        return []
    else:
        raise dash.exceptions.PreventUpdate


@app.callback(
    [
        Output('table-container', 'children'),
        Output('line-chart', 'figure'),
        Output('bar-chart', 'figure')
    ],
    [
        Input('pillar-dropdown', 'value'),
        Input('start-year-dropdown', 'value'),
        Input('end-year-dropdown', 'value'),
        Input('province-checkboxes', 'value'),
        Input('bar-pillar-dropdown', 'value'),
        Input('bar-year-dropdown', 'value')
    ]
)
def update_data(pillar, start_year, end_year, selected_provinces, bar_chart_pillar, bar_chart_year):

    selected_data = pillar_data[pillar]
    filtered_provinces = []
    filtered_scores = []
    filtered_distances_km = []
    filtered_distances_mi = []
    filtered_categories = [] 

    color_palette = px.colors.qualitative.Plotly
    for index, (province, scores, distance_km, distance_mi, category) in enumerate(zip(selected_data['provinces'],
                                                                                       selected_data['scores'],
                                                                                       selected_data['distances_km'],
                                                                                       selected_data['distances_mi'],
                                                                                       selected_data['categories'])):
        if province in selected_provinces:
            filtered_provinces.append(province)
            filtered_scores.append(scores)
            filtered_distances_km.append(distance_km)
            filtered_distances_mi.append(distance_mi)
            filtered_categories.append(category)  

    table_rows = [
        html.Tr([
            html.Th('Province/LGU'),
            html.Th('Distance (km)'),
            html.Th('Distance (mi)'),
            html.Th('Category')
        ])
    ]
    for province, distance_km, distance_mi, category, color in zip(filtered_provinces, filtered_distances_km,
                                                                   filtered_distances_mi, filtered_categories,
                                                                   color_palette):
        table_rows.append(html.Tr([
            html.Td(province),
            html.Td(distance_km),
            html.Td(distance_mi),
            html.Td(category) 
        ]))

    line_chart_data = []
    for province, scores, color in zip(filtered_provinces, filtered_scores, color_palette):
        line_chart_data.append({
            'x': list(range(start_year, end_year + 1)),
            'y': scores,
            'mode': 'lines',
            'name': province,
            'line': {'color': color}
        })

    # Bar chart
    bar_chart_selected_data = pillar_data[bar_chart_pillar]
    bar_chart_filtered_provinces = []
    bar_chart_filtered_scores = []
    for province, scores in zip(bar_chart_selected_data['provinces'], bar_chart_selected_data['scores']):
        if province in selected_provinces:
            bar_chart_filtered_provinces.append(province)
            bar_chart_filtered_scores.append(scores)

    bar_chart_overall_scores = [score[bar_chart_year - start_year] for score in bar_chart_filtered_scores]  # Fixed indexing here
    bar_chart_data = [{
        'x': bar_chart_filtered_provinces,
        'y': bar_chart_overall_scores,
        'type': 'bar',
        'marker': {'color': color_palette}
    }]

    bar_chart_layout = {
        'title': 'Composition of Overall score',
        'xaxis': {'title': 'LGU'},
        'yaxis': {'title': 'Overall Score'}
    }

    return html.Table(table_rows), {'data': line_chart_data, 'layout': {'title': f'{pillar} scores by LGU over Time',
                                                                         'xaxis': {'title': 'Year'},
                                                                         'yaxis': {'title': 'Score'}}}, {
               'data': bar_chart_data, 'layout': bar_chart_layout}

@app.callback(
    Output('pillar-info-container', 'children'),
    [
        Input('bar-pillar-dropdown', 'value'),
        Input('bar-year-dropdown', 'value'),
        Input('province-checkboxes', 'value')
    ]
)
def update_pillar_info(bar_chart_pillar, bar_chart_year, selected_provinces):
    if bar_chart_pillar in pillar_descriptions:
        description = pillar_descriptions[bar_chart_pillar]['Description']

        selected_data = pillar_data[bar_chart_pillar]
        selected_scores = [score[bar_chart_year - 2014] for score in selected_data['scores']]

        table_rows = []

        #Table for LGU and Score based on Bar Chart
        for province, score in zip(selected_data['provinces'], selected_scores):
            if province in selected_provinces:
                if isinstance(score, (int, float)):
                    table_rows.append(html.Tr([
                        html.Td(province),
                        html.Td(str(score)) 
                    ]))
                else:
                    table_rows.append(html.Tr([
                        html.Td(province),
                        html.Td("Data Unavailable")
                    ]))

        #LGU and Score Table
        info_table = html.Table([
            html.Tr([
                html.Th('Description')
            ]),
            html.Tr([
                html.Td(description)
            ])
        ] + table_rows)

        return info_table
    else:
        return 'No information available for selected pillar'


if __name__ == '__main__':
    app.run_server(debug=True)
