import dash
import plotly.express as px
import dash_bootstrap_components as dbc
from dash import dcc, html, Input, Output
from openpyxl import load_workbook
from pathlib import Path

app = dash.Dash(__name__, external_stylesheets=[dbc.themes.BOOTSTRAP])

page1_layout = html.Div([
    html.H1('CMCI HUB'),
    html.P('Explore CMCI Data with Ease')
])

# Load datasets
dataset_folder = Path('Datasets/')
workbook_LGU = load_workbook(dataset_folder / 'LGU_Data/LGUs.xlsx')

pillar_data_LGU = {}
pillar_data_PROV = {}

for sheet in workbook_LGU:
    LGUs = []
    scores = []
    distances_km = []
    distances_mi = []
    categories = [] 
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if len(row) >= 13:  
            LGUs.append(row[0])
            scores.append(row[1:11])
            distances_km.append(row[11])
            distances_mi.append(row[12])
            categories.append(row[13])  

    pillar_name = sheet.title
    pillar_data_LGU[pillar_name] = {
        'LGUs': LGUs,
        'scores': scores,
        'distances_km': distances_km,
        'distances_mi': distances_mi,
        'categories': categories 
    }

workbook_PROV = load_workbook(dataset_folder / 'Province_Data/Prov Dataset.xlsx')

for sheet in workbook_PROV:
    provinces = []
    scores = []
    distances_km = []
    distances_mi = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if len(row) >= 12:  
            provinces.append(row[0])
            scores.append(row[3:12])  
            distances_km.append(row[1])
            distances_mi.append(row[2])

    pillar_name = sheet.title
    pillar_data_PROV[pillar_name] = {
        'provinces': provinces,
        'scores': scores,
        'distances_km': distances_km,
        'distances_mi': distances_mi
    }

pillar_names = list(pillar_data_LGU.keys())
pillar_names = sorted(pillar_names)

all_years = list(range(2014, 2024))

LGUs = list(set(LGU for pillar in pillar_names for LGU in pillar_data_LGU[pillar]['LGUs']))

pillar_descriptions = {
    'overall score': {
        'Description': 'The sum of scores on five main pillars which pool data from several sub-indicators. The higher the score of a city or municipality, the more competitive it is.'
    },
    'economic dynamism': {
        'Description': 'Refers to stable expansion of businesses and industries and higher employment. Matches output and productivity of the local economy with the local resources. Localities are centers of economic activities, and due to this, business expansion and job creation are easily observable in local settings.'
    },
    'government efficiency': {
        'Description': 'Refers to the quality and reliability of government services and government support for effective and sustainable productive expansion. This factor looks at government as an institution that is generally not corrupt; able to protect and enforce contracts; apply moderate and reasonable taxation and is able to regulate proactively.'
    },
    'infrastructure': {
        'Description': 'Pertains to the physical assets that connect, expand, and sustain a locality and its surroundings to enable provision of goods and services. It involves basic inputs of production such as energy, water; interconnection of production such as transportation, roads and communications; sustenance of production such as waste, disaster preparedness, environmental sustainability; and human capital formation infrastructure.'
    },
    'resiliency': {
        'Description': 'Applies to the capacity of a locality to build systems that can absorb change and disturbance and being able to adapt to such changes. It spans frameworks that bind LGUs and their constituents to prepare for possible shocks and stresses; budgeting for disaster risk reduction; hazard/risk identification mechanisms; resilience-related infrastructure; and resilience-related mechanisms.'
    },
    'innovation': {
        'Description': 'Refers to the ability of a locality to harness its creative potential to improve or sustain current levels of productivity. It hinges mainly on the development of creative capital which are human resources, research capabilities, and networking capacities.'
    }
}

# Define the app layout
app.layout = dbc.Container([
    # Header
    html.Div([
        html.H1("DASHBOARD", style={'text-align': 'left', 'font-size': '50px'}),
        html.Div([
            # Level Dropdown
            html.Label('Level'),
            dcc.Dropdown(
                id='level-dropdown',
                options=[
                    {'label': 'LGU', 'value': 'LGU'},
                    {'label': 'Province', 'value': 'Province'}
                ],
                value='LGU',  
                style={'width': '100px', 'display': 'inline-block'}
            ),
            # Pillar Dropdown
            html.Label('Pillar'),
            dcc.Dropdown(
                id='pillar-dropdown',
                options=[{'label': pillar, 'value': pillar} for pillar in pillar_names],
                value=pillar_names[0],  
                style={'width': '150px', 'display': 'inline-block'}
            ),
            # Starting Year Dropdown
            html.Label('Starting Year'),
            dcc.Dropdown(
                id='start-year-dropdown',
                options=[{'label': str(year), 'value': year} for year in all_years],
                value=2014,  
                style={'width': '80px', 'display': 'inline-block'}
            ),
            # End Year Dropdown
            html.Label('End Year'),
            dcc.Dropdown(
                id='end-year-dropdown',
                options=[{'label': str(year), 'value': year} for year in all_years],
                value=2023,  
                style={'width': '80px', 'display': 'inline-block'}
            ),
        ], style={'display': 'flex', 'flex-direction': 'row', 'gap': '20px', 'padding': '20px', 'width':'100%','margin-right':'20px'}),
    ], id='header', style={'display': 'flex', 'flex-direction': 'row', 'text-align': 'center'}),

    # Row 2 (Displayed conditionally)
    dbc.Row([
        # First column 
        dbc.Col([
            html.Div([
                html.H3('LGU Information', style={'text-align': 'center', 'margin-bottom': '10px'}),
                html.Div(id='table-container', style={'margin-bottom': '20px'})
            ])
        ], id='row2-col1', width=3),
        # Second column 
        dbc.Col([
            html.Div([
                html.H3('Scores over Time', style={'text-align': 'center'}),
                dcc.Graph(id='line-chart')
            ])
        ], id='row2-col2', width=6),
        # Third column 
        dbc.Col([
            html.Div([
                html.Label('Select LGUs'),
                dcc.Input(id='LGU-search', type='text', placeholder='Search LGUs...'),
                dcc.Checklist(
                    id='LGU-checkboxes',
                    options=[{'label': LGU, 'value': LGU} for LGU in LGUs],
                    value=[],
                    style={'overflowY': 'scroll', 'height': '400px'}
                ),
                html.Button('Clear Selection', id='clear-selection-button', n_clicks=0)
            ], style={'margin-left': '20px'})
        ], id='row2-col3', width=3)
    ], id='row2', style={'display': 'none'}),

    # Row 3 (Displayed conditionally)
    dbc.Row([
        # First column 
        dbc.Col([
            html.Div(id='pillar-info-container')
        ], id='row3-col1', width=3),
        # Second column 
        dbc.Col([
            html.Div([
                html.H3('Composition of Overall Score', style={'text-align': 'center'}),
                html.Div([
                    html.Label('Pillar', style={'margin-right': '10px'}),
                    dcc.Dropdown(
                        id='bar-pillar-dropdown',
                        options=[{'label': pillar, 'value': pillar} for pillar in pillar_names],
                        value=pillar_names[0],
                        style={'width': '150px', 'margin-bottom': '10px'}
                    ),
                    html.Label('Year', style={'margin-left': '20px', 'margin-right': '10px'}),
                    dcc.Dropdown(
                        id='bar-year-dropdown',
                        options=[{'label': str(year), 'value': year} for year in all_years],
                        value=2023,
                        style={'width': '80px', 'margin-bottom': '10px'}
                    ),
                ], style={'display': 'flex', 'align-items': 'center', 'justify-content': 'center', 'margin-bottom': '20px'}),
                dcc.Graph(id='bar-chart')
            ]),
        ], id='row3-col2', width=6),
        # Third column 
        dbc.Col([], id='row3-col3', width=3)
    ], id='row3', style={'display': 'none'}),

    # Row 4 (Displayed conditionally)
    dbc.Row([html.Div([
        html.Div([
    # Sidebar
    html.Div([
             # Search bar 
        html.Label('Select Provinces',style={'fontWeight': 'bold', 'margin-right':'40px'}),
        dcc.Input(id='province-search', type='text', placeholder='Search Provinces...'),
        dcc.Checklist(
            id='province-checkboxes',
            options=[{'label': province, 'value': province} for province in provinces],
            value=[],
            style={'overflowY': 'scroll', 'height': '1050px', 'font-size':'20px', 'margin-right':'2%'}
        ),
        html.Button('Clear Selection', id='clear-selection-button-prov', n_clicks=0)
    ], style={'width': '20%', 'float': 'right', 'margin-left': '2px','border': '1px solid #ccc','font-size':'20px'}),
]),

    # Line chart 
    html.Div([
        html.H3('CMCI Scores', style={'text-align': 'center'}),
        dcc.Graph(id='line-chart-prov'),
    ], style={'width': '48%', 'height': '50vh', 'display': 'inline-block', 'margin-right': '2%', 'margin-top': '2%'}),

  # Bar chart
    html.Div([
        html.H3('Distance of Each Province to the Center of Manila', style={'text-align': 'center'}),
        dcc.Graph(id='bar-chart-prov')
    ], style={'width': '48%', 'height': '50vh', 'display': 'inline-block', 'margin-right': '2%'}),
    ])
    ], id='row4', style={'display': 'none'})
], fluid=True)

@app.callback(
    [Output('row2', 'style'), Output('row3', 'style'), Output('row4', 'style')],
    [Input('level-dropdown', 'value')]
)
def update_row_visibility(level):
    row2_style = {'display': 'none'} if level != 'LGU' else {}
    row3_style = {'display': 'none'} if level != 'LGU' else {}
    row4_style = {'display': 'none'} if level != 'Province' else {}
    return row2_style, row3_style, row4_style


@app.callback(
    Output('province-checkboxes', 'options'),
    [Input('province-search', 'value')]
)
def update_province_options_prov(search_value):
    if search_value is None:
        return [{'label': province, 'value': province} for province in provinces]
    else:
        filtered_provinces = [province for province in provinces if search_value.lower() in province.lower()]
        return [{'label': province, 'value': province} for province in filtered_provinces]

# Callback to clear selected provinces 
@app.callback(
    Output('province-checkboxes', 'value'),
    [Input('clear-selection-button-prov', 'n_clicks')]
)
def clear_selected_provinces(n_clicks):
    if n_clicks > 0:
        return []
    else:
        raise dash.exceptions.PreventUpdate

# Callback for line chart
@app.callback(
    Output('line-chart-prov', 'figure'),
    [
        Input('pillar-dropdown', 'value'),
        Input('start-year-dropdown', 'value'),
        Input('end-year-dropdown', 'value'),
        Input('province-checkboxes', 'value'),
    ]
)
def update_data_prov(pillar, start_year, end_year, selected_provinces):
    # Filter data based on selected pillar and years
    selected_data = pillar_data_PROV[pillar]
    filtered_provinces = []
    filtered_scores = []
   
    # Colors 
    color_palette = px.colors.qualitative.Plotly
    for index, (province, scores) in enumerate(zip(selected_data['provinces'], selected_data['scores'])):
        if province in selected_provinces:
            filtered_provinces.append(province)
            filtered_scores.append(scores)

    line_chart_data = []
    for province, scores, color in zip(filtered_provinces, filtered_scores, color_palette):
        line_chart_data.append({
            'x': list(range(start_year - 1, end_year)),
            'y': scores,
            'mode': 'lines',
            'name': province,
            'line': {'color': color}
        })

    return {'data': line_chart_data, 'layout': {'title': f'{pillar} scores by Province over Time',
                                                'xaxis': {'title': 'Year'},
                                                'yaxis': {'title': 'Score'}}}

@app.callback(
    Output('bar-chart-prov', 'figure'),
    [
        Input('pillar-dropdown', 'value'),
        Input('start-year-dropdown', 'value'),
        Input('end-year-dropdown', 'value'),
        Input('province-checkboxes', 'value')
    ]
)
def update_bar_chart_prov(pillar, start_year, end_year, selected_provinces):
    # Bar chart data
    bar_chart_selected_data = pillar_data_PROV[pillar]
    bar_chart_filtered_provinces = []
    bar_chart_filtered_distances_mi = []
    for province, distance_mi in zip(bar_chart_selected_data['provinces'], bar_chart_selected_data['distances_mi']):
        if province in selected_provinces:
            bar_chart_filtered_provinces.append(province)
            bar_chart_filtered_distances_mi.append(distance_mi)

    bar_chart_data = [{
        'x': bar_chart_filtered_provinces,
        'y': bar_chart_filtered_distances_mi,
        'type': 'bar',
        'marker': {'color': px.colors.qualitative.Plotly} 
    }]

    # Bar chart layout
    bar_chart_layout = {
        'title': 'Distance (in mi)',
        'xaxis': {'title': 'Province'},
        'yaxis': {'title': 'Distance (in mi)'}
    }

    return {'data': bar_chart_data, 'layout': bar_chart_layout}

@app.callback(
    Output('LGU-checkboxes', 'options'),
    [Input('LGU-search', 'value')]
)
def update_LGU_options(search_value):
    if search_value is None:
        sorted_LGUs = sorted(LGUs)
        return [{'label': LGU, 'value': LGU} for LGU in sorted_LGUs]
    else:
        filtered_LGUs = [LGU for LGU in LGUs if search_value.lower() in LGU.lower()]
        sorted_filtered_LGUs = sorted(filtered_LGUs)
        return [{'label': LGU, 'value': LGU} for LGU in sorted_filtered_LGUs]


@app.callback(
    Output('LGU-checkboxes', 'value'),
    [Input('clear-selection-button', 'n_clicks')]
)
def clear_selected_LGUs(n_clicks):
    if n_clicks > 0:
        return []
    else:
        raise dash.exceptions.PreventUpdate


@app.callback(
    [
        Output('table-container', 'children'),
        Output('line-chart', 'figure'),
        Output('bar-chart', 'figure')
    ],
    [
        Input('pillar-dropdown', 'value'),
        Input('start-year-dropdown', 'value'),
        Input('end-year-dropdown', 'value'),
        Input('LGU-checkboxes', 'value'),
        Input('bar-pillar-dropdown', 'value'),
        Input('bar-year-dropdown', 'value')
    ]
)
def update_data(pillar, start_year, end_year, selected_LGUs, bar_chart_pillar, bar_chart_year):

    selected_data = pillar_data_LGU[pillar]
    filtered_LGUs = []
    filtered_scores = []
    filtered_distances_km = []
    filtered_distances_mi = []
    filtered_categories = [] 

    color_palette = px.colors.qualitative.Plotly
    for index, (LGU, scores, distance_km, distance_mi, category) in enumerate(zip(selected_data['LGUs'],
                                                                                       selected_data['scores'],
                                                                                       selected_data['distances_km'],
                                                                                       selected_data['distances_mi'],
                                                                                       selected_data['categories'])):
        if LGU in selected_LGUs:
            filtered_LGUs.append(LGU)
            filtered_scores.append(scores)
            filtered_distances_km.append(distance_km)
            filtered_distances_mi.append(distance_mi)
            filtered_categories.append(category)  

    table_rows = [
        html.Tr([
            html.Th('LGU'),
            html.Th('Distance (km)'),
            html.Th('Distance (mi)'),
            html.Th('Category')
        ])
    ]
    for LGU, distance_km, distance_mi, category, color in zip(filtered_LGUs, filtered_distances_km,
                                                                   filtered_distances_mi, filtered_categories,
                                                                   color_palette):
        table_rows.append(html.Tr([
            html.Td(LGU),
            html.Td(distance_km),
            html.Td(distance_mi),
            html.Td(category) 
        ]))

    line_chart_data = []
    for LGU, scores, color in zip(filtered_LGUs, filtered_scores, color_palette):
        line_chart_data.append({
            'x': list(range(start_year, end_year + 1)),
            'y': scores,
            'mode': 'lines',
            'name': LGU,
            'line': {'color': color}
        })

    # Bar chart
    bar_chart_selected_data = pillar_data_LGU[bar_chart_pillar]
    bar_chart_filtered_LGUs = []
    bar_chart_filtered_scores = []
    for LGU, scores in zip(bar_chart_selected_data['LGUs'], bar_chart_selected_data['scores']):
        if LGU in selected_LGUs:
            bar_chart_filtered_LGUs.append(LGU)
            bar_chart_filtered_scores.append(scores)

    bar_chart_overall_scores = [score[bar_chart_year - start_year] for score in bar_chart_filtered_scores]  
    bar_chart_data = [{
        'x': bar_chart_filtered_LGUs,
        'y': bar_chart_overall_scores,
        'type': 'bar',
        'marker': {'color': color_palette}
    }]

    bar_chart_layout = {
        'title': 'Composition of Overall score',
        'xaxis': {'title': 'LGU'},
        'yaxis': {'title': 'Overall Score'}
    }

    return html.Table(table_rows), {'data': line_chart_data, 'layout': {'title': f'{pillar} scores by LGU over Time',
                                                                         'xaxis': {'title': 'Year'},
                                                                         'yaxis': {'title': 'Score'}}}, {
               'data': bar_chart_data, 'layout': bar_chart_layout}

@app.callback(
    Output('pillar-info-container', 'children'),
    [
        Input('bar-pillar-dropdown', 'value'),
        Input('bar-year-dropdown', 'value'),
        Input('LGU-checkboxes', 'value')
    ]
)
def update_pillar_info(bar_chart_pillar, bar_chart_year, selected_LGUs):
    if bar_chart_pillar in pillar_descriptions:
        description = pillar_descriptions[bar_chart_pillar]['Description']

        selected_data = pillar_data_LGU[bar_chart_pillar]
        selected_scores = [score[bar_chart_year - 2014] for score in selected_data['scores']]

        table_rows = []

        # Table for LGU and Score based on Bar Chart
        for LGU, score in zip(selected_data['LGUs'], selected_scores):
            if LGU in selected_LGUs:
                if isinstance(score, (int, float)):
                    table_rows.append(html.Tr([
                        html.Td(LGU),
                        html.Td(str(score)) 
                    ]))
                else:
                    table_rows.append(html.Tr([
                        html.Td(LGU),
                        html.Td("Data Unavailable")
                    ]))

         # LGU and Score Table
        info_table = html.Table([
        html.Tr([
            html.Th(html.H3('Pillar Description'), style={'text-align': 'center', 'font-weight': 'bold'})  
        ]),
        html.Tr([
            html.Td(html.B(bar_chart_pillar.upper()), style={'font-size': 'small', 'font-weight': 'bold'})
        ]),
        html.Tr([
            html.Td(description, style={'text-align': 'justify'}) 
        ])
    ] + table_rows)

        return info_table
    else:
        return 'No information available for selected pillar'


page3_layout = html.Div([
    html.H1('INTERACTIVE MAP'),
    html.P('This is page 3.')
])

# Define navigation bar
navbar = dbc.NavbarSimple(
    children=[
        html.Div(
            [
                dbc.Button("🌐 CMCI HUB", href="/page-1", color="secondary", className="me-3", style={'text-align':'left'}),
                dbc.Button("📊 VISUALIZATION DASHBOARD", href="/page-2", color="secondary", className="me-3", style={''}),
                dbc.Button("🗾 INTERACTIVE MAP", href="/page-3", color="secondary")
            ],
            className="d-flex justify-content-center"
        )
    ],
    color="dark", 
    dark=True,  
    style={"font-family": "Arial, sans-serif", "font-weight": "bold", "color": "black"}  # Apply font style
)

@app.callback(Output('page-content', 'children'),
              [Input('url', 'pathname')])
def display_page(pathname):
    if pathname == '/' or pathname == '/page-1':
        return page1_layout
    elif pathname == '/page-2':
        return app.layout
    elif pathname == '/page-3':
        return page3_layout
    else:
        return '404 - Page not found'


if __name__ == '__main__':
    app.run_server(debug=True)

