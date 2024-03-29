import dash
import plotly.express as px
from dash import dcc, html, Input, Output
from openpyxl import load_workbook
import dash_bootstrap_components as dbc

workbook = load_workbook(r'C:\Users\ASUS\OneDrive\Documents\Desktop\test\Datasets\Province_Data\Prov Dataset.xlsx')

pillar_data = {}

for sheet in workbook:
    provinces = []
    scores = []
    distances_km = []
    distances_mi = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if len(row) >= 12:  
            provinces.append(row[0])
            scores.append(row[3:12])  
            distances_km.append(row[1])
            distances_mi.append(row[2])

    pillar_name = sheet.title
    pillar_data[pillar_name] = {
        'provinces': provinces,
        'scores': scores,
        'distances_km': distances_km,
        'distances_mi': distances_mi
    }

pillar_names = list(pillar_data.keys())

all_years = list(range(2014, 2024))

provinces = list(set(province for pillar in pillar_names for province in pillar_data[pillar]['provinces']))


app = dash.Dash(__name__, external_stylesheets=[dbc.themes.BOOTSTRAP])

page1_layout = html.Div([
    html.H1('CMCI HUB'),
    html.P('Explore CMCI Data with Ease')
])

page2_layout = html.Div([
    # Header
    html.Div([
        html.H1("DASHBOARD", style={'text-align': 'left','font-size':'50px'}),
        html.Div([
            # Level Dropdown
            html.Label('Level'),
            dcc.Dropdown(
                id='level-dropdown',
                options=[
                    {'label': 'Province', 'value': 'Province'},
                    {'label': 'LGU', 'value': 'LGU'}
                ],
                value='Province',  
                style={'width': '100px', 'display': 'inline-block'}
            ),
            # Pillar Dropdown
            html.Label('Pillar'),
            dcc.Dropdown(
                id='pillar-dropdown',
                options=[{'label': pillar, 'value': pillar} for pillar in pillar_names],
                value=pillar_names[0],  
                style={'width': '150px', 'display': 'inline-block'}
            ),
            # Starting Year Dropdown
            html.Label('Starting Year'),
            dcc.Dropdown(
                id='start-year-dropdown',
                options=[{'label': str(year), 'value': year} for year in all_years],
                value=2014,  
                style={'width': '80px', 'display': 'inline-block'}
            ),
            # End Year Dropdown
            html.Label('End Year'),
            dcc.Dropdown(
                id='end-year-dropdown',
                options=[{'label': str(year), 'value': year} for year in all_years],
                value=2023,  
                style={'width': '80px', 'display': 'inline-block'}
            ),
        ], style={'display': 'flex', 'flex-direction': 'row', 'gap': '20px', 'padding': '20px', 'width':'100%','margin-right':'20px'}),
    ], style={'display': 'flex', 'flex-direction': 'row','text-align': 'center'}),

    html.Div([
        html.Div([
    # Sidebar
    html.Div([
             # Search bar 
        html.Label('Select Provinces',style={'fontWeight': 'bold'}),
        dcc.Input(id='province-search', type='text', placeholder='Search Provinces...'),
        dcc.Checklist(
            id='province-checkboxes',
            options=[{'label': province, 'value': province} for province in provinces],
            value=[],
            style={'overflowY': 'scroll', 'height': '400px'}
        ),
        html.Button('Clear Selection', id='clear-selection-button', n_clicks=0, style={'fontSize':'20px'})
    ], style={'width': '20%', 'float': 'right', 'margin-left': '10px','border': '1px solid #ccc','font-size':'25px'}),
]),
        # Table province data
        html.Div(id='table-container', style={'width': '30%', 'float': 'left', 'border': '1px solid #ccc',}),

    # Line chart 
    html.Div([
        html.H3('CMCI Scores', style={'text-align': 'center'}),
        dcc.Graph(id='line-chart'),
    ], style={'width': '48%', 'height': '50vh', 'display': 'inline-block', 'margin-right': '2%', 'margin-top': '2%'}),

  # Bar chart
    html.Div([
        html.H3('Distance of Each Province to the Center of Manila', style={'text-align': 'center'}),
        dcc.Graph(id='bar-chart')
    ], style={'width': '48%', 'height': '50vh', 'display': 'inline-block', 'margin-right': '2%'}),
    ])
])

@app.callback(
    Output('province-checkboxes', 'options'),
    [Input('province-search', 'value')]
)
def update_province_options(search_value):
    if search_value is None:
        return [{'label': province, 'value': province} for province in provinces]
    else:
        filtered_provinces = [province for province in provinces if search_value.lower() in province.lower()]
        return [{'label': province, 'value': province} for province in filtered_provinces]

# Callback to clear selected provinces 
@app.callback(
    Output('province-checkboxes', 'value'),
    [Input('clear-selection-button', 'n_clicks')]
)
def clear_selected_provinces(n_clicks):
    if n_clicks > 0:
        return []
    else:
        raise dash.exceptions.PreventUpdate

# Callback for table and line chart
@app.callback(
    [
        Output('table-container', 'children'),
        Output('line-chart', 'figure'),
    ],
    [
        Input('pillar-dropdown', 'value'),
        Input('start-year-dropdown', 'value'),
        Input('end-year-dropdown', 'value'),
        Input('province-checkboxes', 'value'),
    ]
)
def update_data(pillar, start_year, end_year, selected_provinces):
    # Filter data based on selected pillar and years
    selected_data = pillar_data[pillar]
    filtered_provinces = []
    filtered_scores = []
    filtered_distances_km = []
    filtered_distances_mi = []
    # Colors 
    color_palette = px.colors.qualitative.Plotly
    for index, (province, scores, distance_km, distance_mi) in enumerate(zip(selected_data['provinces'],
                                                                              selected_data['scores'],
                                                                              selected_data['distances_km'],
                                                                              selected_data['distances_mi'])):
        if province in selected_provinces:
            filtered_provinces.append(province)
            filtered_scores.append(scores)
            filtered_distances_km.append(distance_km)
            filtered_distances_mi.append(distance_mi)

    table_rows = [
        html.Tr([
            html.Th('Province/LGU'),
            html.Th('Distance (km)'),
            html.Th('Distance (mi)')
        ])
    ]
    for province, distance_km, distance_mi, color in zip(filtered_provinces, filtered_distances_km, filtered_distances_mi, color_palette):
        table_rows.append(html.Tr([
            html.Td(province),
            html.Td(distance_km),
            html.Td(distance_mi)
        ]))

    line_chart_data = []
    for province, scores, color in zip(filtered_provinces, filtered_scores, color_palette):
        line_chart_data.append({
            'x': list(range(start_year, end_year + 1)),
            'y': scores,
            'mode': 'lines',
            'name': province,
            'line': {'color': color}
        })

    return html.Table(table_rows), {'data': line_chart_data, 'layout': {'title': f'{pillar} scores by Province over Time',
                                                                         'xaxis': {'title': 'Year'},
                                                                         'yaxis': {'title': 'Score'}}}

@app.callback(
    Output('bar-chart', 'figure'),
    [
        Input('pillar-dropdown', 'value'),
        Input('start-year-dropdown', 'value'),
        Input('end-year-dropdown', 'value'),
        Input('province-checkboxes', 'value')
    ]
)
def update_bar_chart(pillar, start_year, end_year, selected_provinces):
    # Bar chart data
    bar_chart_selected_data = pillar_data[pillar]
    bar_chart_filtered_provinces = []
    bar_chart_filtered_distances_mi = []
    for province, distance_mi in zip(bar_chart_selected_data['provinces'], bar_chart_selected_data['distances_mi']):
        if province in selected_provinces:
            bar_chart_filtered_provinces.append(province)
            bar_chart_filtered_distances_mi.append(distance_mi)

    bar_chart_data = [{
        'x': bar_chart_filtered_provinces,
        'y': bar_chart_filtered_distances_mi,
        'type': 'bar',
        'marker': {'color': px.colors.qualitative.Plotly} 
    }]

    # Bar chart layout
    bar_chart_layout = {
        'title': 'Distance (in mi)',
        'xaxis': {'title': 'Province'},
        'yaxis': {'title': 'Distance (in mi)'}
    }

    return {'data': bar_chart_data, 'layout': bar_chart_layout}

# Map w/Distance (Placholder) 


page3_layout = html.Div([
    html.H1('INTERACTIVE MAP'),
    html.P('This is page 3.')
])

# Define navigation bar
navbar = dbc.NavbarSimple(
    children=[
        html.Div(
            [
                dbc.Button("CMCI HUB", href="/page-1", color="secondary", className="me-2"),
                dbc.Button("VISUALIZATION DASHBOARD", href="/page-2", color="secondary", className="me-2"),
                dbc.Button("INTERACTIVE MAP", href="/page-3", color="secondary")
            ],
            className="d-flex justify-content-center"
        )
    ],
    color="dark", 
    dark=True,  
    style={"font-family": "Arial, sans-serif", "font-weight": "bold", "color": "black"}  # Apply font style
)

@app.callback(Output('page-content', 'children'),
              [Input('url', 'pathname')])
def display_page(pathname):
    if pathname == '/' or pathname == '/page-1':
        return page1_layout
    elif pathname == '/page-2':
        return page2_layout
    elif pathname == '/page-3':
        return page3_layout
    else:
        return '404 - Page not found'

app.layout = html.Div([
    dcc.Location(id='url', refresh=False),
    navbar,
    html.Div(id='page-content')
])

if __name__ == '__main__':
    app.run_server(debug=True)
