import dash
from dash import Dash, html, dcc, callback, Output, Input
import plotly.express as px
import dash_bootstrap_components as dbc
import pandas as pd
import geopandas as gpd
import numpy as np
from openpyxl import load_workbook
from pathlib import Path
import plotly.graph_objects as go
import numpy as np

dataset_folder = Path('Datasets/')
workbook = load_workbook(dataset_folder / 'InteractiveMap_Data/InteractiveMap_Profile.xlsx')

all_years = list(range(2014, 2024))

# Map
file_paths = [
    "C:/Users/acer/Documents/Data Sci/DataSets/Province JSON/provinces-region-bicolregionregionv.json",
    "C:/Users/acer/Documents/Data Sci/DataSets/Province JSON/provinces-region-autonomousregionofmuslimmindanaoarmm.json",
    "C:/Users/acer/Documents/Data Sci/DataSets/Province JSON/provinces-region-cagayanvalleyregionii.json",
    "C:/Users/acer/Documents/Data Sci/DataSets/Province JSON/provinces-region-calabarzonregioniva.json",
    "C:/Users/acer/Documents/Data Sci/DataSets/Province JSON/provinces-region-caragaregionxiii.json",
    "C:/Users/acer/Documents/Data Sci/DataSets/Province JSON/provinces-region-centralluzonregioniii.json",
    "C:/Users/acer/Documents/Data Sci/DataSets/Province JSON/provinces-region-centralvisayasregionvii.json",
    "C:/Users/acer/Documents/Data Sci/DataSets/Province JSON/provinces-region-cordilleraadministrativeregioncar.json",
    "C:/Users/acer/Documents/Data Sci/DataSets/Province JSON/provinces-region-davaoregionregionxi.json",
    "C:/Users/acer/Documents/Data Sci/DataSets/Province JSON/provinces-region-easternvisayasregionviii.json",
    "C:/Users/acer/Documents/Data Sci/DataSets/Province JSON/provinces-region-ilocosregionregioni.json",
    "C:/Users/acer/Documents/Data Sci/DataSets/Province JSON/provinces-region-metropolitanmanila.json",
    "C:/Users/acer/Documents/Data Sci/DataSets/Province JSON/provinces-region-mimaroparegionivb.json",
    "C:/Users/acer/Documents/Data Sci/DataSets/Province JSON/provinces-region-northernmindanaoregionx.json",
    "C:/Users/acer/Documents/Data Sci/DataSets/Province JSON/provinces-region-soccsksargenregionxii.json",
    "C:/Users/acer/Documents/Data Sci/DataSets/Province JSON/provinces-region-westernvisayasregionvi.json",
    "C:/Users/acer/Documents/Data Sci/DataSets/Province JSON/provinces-region-zamboangapeninsularegionix.json"
]

# Read each file and append to a list
ph_list = [gpd.read_file(file) for file in file_paths]

# Combine all GeoDataFrames in the list into a single GeoDataFrame
ph = gpd.GeoDataFrame(pd.concat(ph_list, ignore_index=True), crs=ph_list[0].crs)
#ph = ph.to_crs(epsg=32651)
p_score = pd.read_csv(dataset_folder / 'Province_Data/Overall Score.csv', encoding='latin1')
ph.loc[ph['PROVINCE'] == 'Agusan del Norte', 'PROVINCE'] = 'Agusan Del Norte'
ph.loc[ph['PROVINCE'] == 'Agusan del Sur', 'PROVINCE'] = 'Agusan Del Sur'
ph.loc[ph['PROVINCE'] == 'Batangas', 'PROVINCE'] = 'Batangas Province'
ph.loc[ph['PROVINCE'] == 'Biliran', 'PROVINCE'] = 'Biliran Province'
ph.loc[ph['PROVINCE'] == 'Cavite', 'PROVINCE'] = 'Cavite Province'
ph.loc[ph['PROVINCE'] == 'Cebu', 'PROVINCE'] = 'Cebu Province'
ph.loc[ph['PROVINCE'] == 'North Cotabato', 'PROVINCE'] = 'Cotabato (North Cotabato)'
ph.loc[ph['PROVINCE'] == 'Davao de Oro', 'PROVINCE'] = 'Davao De Oro'
ph.loc[ph['PROVINCE'] == 'Davao del Norte', 'PROVINCE'] = 'Davao Del Norte'
ph.loc[ph['PROVINCE'] == 'Davao del Sur', 'PROVINCE'] = 'Davao Del Sur'
ph.loc[ph['PROVINCE'] == 'Iloilo', 'PROVINCE'] = 'Iloilo Province'
ph.loc[ph['PROVINCE'] == 'Lanao del Norte', 'PROVINCE'] = 'Lanao Del Norte'
ph.loc[ph['PROVINCE'] == 'Lanao del Sur', 'PROVINCE'] = 'Lanao Del Sur'
ph.loc[ph['PROVINCE'] == 'Leyte', 'PROVINCE'] = 'Leyte Province'
ph.loc[ph['PROVINCE'] == 'Masbate', 'PROVINCE'] = 'Masbate Province'
ph.loc[ph['PROVINCE'] == 'Romblon', 'PROVINCE'] = 'Romblon Province'
ph.loc[ph['PROVINCE'] == 'Samar', 'PROVINCE'] = 'Samar (Western Samar)'
ph.loc[ph['PROVINCE'] == 'Siquijor', 'PROVINCE'] = 'Siquijor Province'
ph.loc[ph['PROVINCE'] == 'Sorsogon', 'PROVINCE'] = 'Sorsogon Province'
ph.loc[ph['PROVINCE'] == 'Surigao del Norte', 'PROVINCE'] = 'Surigao Del Norte'
ph.loc[ph['PROVINCE'] == 'Surigao del Sur', 'PROVINCE'] = 'Surigao Del Sur'
ph.loc[ph['PROVINCE'] == 'Tarlac', 'PROVINCE'] = 'Tarlac Province'
ph.loc[ph['PROVINCE'] == 'Zamboanga del Norte', 'PROVINCE'] = 'Zamboanga Del Norte'
ph.loc[ph['PROVINCE'] == 'Zamboanga del Sur', 'PROVINCE'] = 'Zamboanga Del Sur'
#ph.loc[ph['PROVINCE'] == 'Maguindanao del Norte', 'PROVINCE'] = 'Maguindanao'
#ph.loc[ph['PROVINCE'] == 'Maguindanao del Sur', 'PROVINCE'] = 'Maguindanao'
ph.loc[ph['PROVINCE'] == 'Metropolitan Manila', 'PROVINCE'] = 'Metro Manila'
p_choro = pd.merge(ph, p_score,  left_on='PROVINCE', right_on='PROVINCE / LGU', how='left', indicator=True)

initial_column_values = p_choro.set_index('PROVINCE')['2023'].replace('-', np.nan).astype(float)
initial_column_values = initial_column_values.fillna(0).astype(int)

initial_fig = px.choropleth(
    p_choro,
    geojson=p_choro.geometry,
    locations=p_choro.index,
    color=initial_column_values,
    color_continuous_scale='Viridis',
    labels={'2023': 'Annual Inflation Rate'},
)

initial_fig.update_geos(fitbounds="locations", visible=False, bgcolor="#C9D1D2")
initial_fig.update_layout(
    coloraxis_colorbar=dict(title='Overall CMCI Score'),
    paper_bgcolor="#C9D1D2",
    geo=dict(
        visible=False,
        bgcolor='rgba(255,255,255,0)'
    ),
)

province_options = [{'label': province, 'value': province} for province in p_choro['PROVINCE'] if province is not None]

# Province profile 
province_sheet = workbook['Province']

region = []
population = []
province_revenue = []
rank = []

for row in province_sheet.iter_rows(min_row=2, values_only=True):
   region.append(row[1])
   population.append(row[2])
   province_revenue.append(row[3])
   rank.append(row[4])

region_data = [row[1].value for row in province_sheet.iter_rows(min_row=2, max_col=2)]
population_data = [row[2].value for row in province_sheet.iter_rows(min_row=2, max_col=3)]
province_revenue_data = [row[3].value for row in province_sheet.iter_rows(min_row=2, max_col=4)]
rank_data = [row[4].value for row in province_sheet.iter_rows(min_row=2, max_col=5)]

def get_province_region(selected_province):
   try:
       index = province_data.index(selected_province)
       return region_data[index]
   except ValueError:
       return 'No data available'
   
def get_province_population(selected_province):
   try:
       index = province_data.index(selected_province)
       return population_data[index]
   except ValueError:
       return 'No data available'


def get_province_revenue(selected_province):
   try:
       index = province_data.index(selected_province)
       return province_revenue[index]
   except ValueError:
       return 'No data available'


def get_province_rank(selected_province):
   try:
       index = province_data.index(selected_province)
       return rank_data[index]
   except ValueError:
       return 'No data available'

# LGU Profile
lgu_sheet = workbook['LGU']

lgu = []
category = []
percentage = []
province = []
revenue = []

for row in lgu_sheet.iter_rows(min_row=2, values_only=True):
   lgu.append(row[0])
   category.append(row[1])
   percentage.append(row[2])
   province.append(row[3])
   revenue.append(row[4])

lgu_data = [row[0].value for row in lgu_sheet.iter_rows(min_row=2, max_col=1)]
category_data = [row[1].value for row in lgu_sheet.iter_rows(min_row=2, max_col=2)]
province_data = [row[3].value for row in lgu_sheet.iter_rows(min_row=2, max_col=4)]
revenue_data = [row[4].value for row in lgu_sheet.iter_rows(min_row=2, max_col=5)]
lgu_options = [{'label': lgu_name, 'value': lgu_name} for lgu_name in lgu_data if lgu_name is not None]

def get_pillar_description(selected_pillar):
   pillar_descriptions = {
       'Resiliency': 'Applies to the capacity of a locality to build systems that can absorb change and disturbance and being able to adapt to such changes. It spans frameworks that bind LGUs and their constituents to prepare for possible shocks and stresses; budgeting for disaster risk reduction; hazard/risk identification mechanisms; resilience-related infrastructure; and resilience-related mechanisms.',
       'Government Efficiency': 'Refers to the quality and reliability of government services and government support for effective and sustainable productive expansion. This factor looks at government as an institution that is generally not corrupt; able to protect and enforce contracts; apply moderate and reasonable taxation and is able to regulate proactively.',
       'Innovation': 'Refers to the ability of a locality to harness its creative potential to improve or sustain current levels of productivity. It hinges mainly on the development of creative capital which are human resources, research capabilities, and networking capacities..',
       'Economic Dynamism': 'Refers to stable expansion of businesses and industries and higher employment. Matches output and productivity of the local economy with the local resources. Localities are centers of economic activities, and due to this, business expansion and job creation are easily observable in local settings.',
       'Infrastructure': 'Pertains to the physical assets that connect, expand, and sustain a locality and its surroundings to enable provision of goods and services. It involves basic inputs of production such as energy, water; interconnection of production such as transportation, roads and communications; sustenance of production such as waste, disaster preparedness, environmental sustainability; and human capital formation infrastructure.'
   }
   return pillar_descriptions.get(selected_pillar, 'No description available')

def get_lgu_province(selected_lgu):
   try:
       index = lgu_data.index(selected_lgu)
       return province_data[index]
   except ValueError:
       return 'No data available'

def get_lgu_category(selected_lgu):
   try:
       index = lgu_data.index(selected_lgu)
       return category_data[index]
   except ValueError:
       return 'No data available'

def get_lgu_revenue(selected_lgu):
   try:
       index = lgu_data.index(selected_lgu)
       return revenue_data[index]
   except ValueError:
       return 'No data available'

app = Dash(__name__, external_stylesheets=[dbc.themes.BOOTSTRAP])

app.layout = dbc.Container([
  
   dbc.Row([
       
   # First column
   dbc.Col([
       dbc.Row([
       dcc.Dropdown(
                                id='map-year-dropdown-province',
                                options=[{'label': str(year), 'value': year} for year in all_years],
                                value=2023,
                                style={'width': '80px', 'margin-bottom': '1px', 'display': 'inline-block', 'vertical-align': 'middle'}
                            ),
       ]),
       dbc.Row([
        html.Div([
        html.H3('Choropleth Map', style={'text-align': 'center', 'margin-bottom': '10px'}),
        dcc.Graph(id='choropleth-map', figure=initial_fig)  # Display the initial choropleth map
    ])
       ]),
       
   ]),
  
   # Second column
   dbc.Col([
       html.Div([
           html.H3('Province Profile', style={'text-align': 'center'}),
           html.Label('Select Province'),
           dcc.Dropdown(
               id='province-dropdown',
               options=province_options,
               value=[]
           ),
           dbc.Row([
               dbc.Col([
                   html.Label('Region'),
                   html.Div(id='region-label')
               ]),
               dbc.Col([
                   html.Label('Population'),
                   html.Div(id='population-label')
               ]),
               dbc.Col([
                   html.Label('Revenue'),
                   html.Div(id='province-revenue-label')
               ]),
               dbc.Col([
                   html.Label('Rank'),
                   html.Div(id='rank-label')
               ])
           ]),
           html.H3('LGU Profile', style={'text-align': 'center'}),

           dbc.Row([
               
           # LGU Dropdown
           html.Label('Select LGU'),
           dcc.Dropdown(
               id='lgu-dropdown',
               options=lgu_options,
               value=[]
           ),
           dbc.Row([
               dbc.Col([
                   html.Label('Select Pillar'),
                   dcc.Dropdown(
                       id='pillar-dropdown',
                       options=[
                           {'label': 'Resiliency', 'value': 'Resiliency'},
                           {'label': 'Government Efficiency', 'value': 'Government Efficiency'},
                           {'label': 'Innovation', 'value': 'Innovation'},
                           {'label': 'Economic Dynamism', 'value': 'Economic Dynamism'},
                           {'label': 'Infrastructure', 'value': 'Infrastructure'},
                       ],
                       value=[]
                   ),
               ]),
               dbc.Col([
                   html.Label('Pillar Description'),
                   html.Div(id='pillar-description')
               ])
           ]),
           dbc.Row([
               dbc.Col([
                   html.Label('Province'),
                   html.Div(id='province-label')
               ]),
               dbc.Col([
                   html.Label('Category'),
                   html.Div(id='category-label')
               ]),
               dbc.Col([
                   html.Label('Revenue'),
                   html.Div(id='revenue-label')
               ])
           ])
       ])
   ]),
   ]),
  
   # Third column
   dbc.Col([
       html.Div([
           html.H3('Score per Pillar', style={'text-align': 'center', 'margin-bottom': '10px'}),
           dcc.Graph(id='bar-chart')
       ])
   ])
   ])
])

# Descriptions
@app.callback(
   [
       Output('pillar-description', 'children'),
       Output('province-label', 'children'),
       Output('category-label', 'children'),
       Output('revenue-label', 'children'),
   ],
   [
       Input('lgu-dropdown', 'value'),
       Input('pillar-dropdown', 'value')
   ]
)
def update_labels(selected_lgu, selected_pillar):
   if selected_lgu and selected_pillar:
       pillar_description = get_pillar_description(selected_pillar)
       lgu_province = get_lgu_province(selected_lgu)
       lgu_category = get_lgu_category(selected_lgu)
       lgu_revenue = get_lgu_revenue(selected_lgu)


       return pillar_description, lgu_province, lgu_category, lgu_revenue
   else:
       return '-', '-', '-', '-'
   
@app.callback(
   [
       Output('region-label', 'children'),
       Output('population-label', 'children'),
       Output('province-revenue-label', 'children'),
       Output('rank-label', 'children'),
   ],
   [
       Input('province-dropdown', 'value')
   ]
)
def update_labels(selected_province):
   if selected_province:
       province_region = get_province_region(selected_province)
       province_population = get_province_population(selected_province)
       province_revenue = get_province_revenue(selected_province)
       province_rank = get_province_rank(selected_province)

       return province_region, province_population, province_revenue, province_rank
   else:
       return '-', '-', '-', '-'


# Bar Chart
@app.callback(
   Output('bar-chart', 'figure'),
   Input('lgu-dropdown', 'value')
)
def update_bar_chart(selected_lgu):
   if not selected_lgu:
       return {}

   lgu_index = lgu_data.index(selected_lgu) + 2
   lgu_data_row = list(lgu_sheet.iter_rows(min_row=lgu_index, max_row=lgu_index, min_col=6, max_col=10, values_only=True))[0]
   pillars = ['Resiliency', 'Government Efficiency', 'Innovation', 'Economic Dynamism', 'Infrastructure']

   fig = px.bar(
       x=pillars,
       y=lgu_data_row,
       labels={'x': 'Pillar', 'y': 'Score'},
       title=f'Scores per Pillar for {selected_lgu}'
   )

   return fig


# Map


if __name__ == '__main__':
   app.run_server(debug=True)

