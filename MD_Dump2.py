import dash
from dash import Dash, html, dcc, callback, Output, Input
import plotly.express as px
import dash_bootstrap_components as dbc
import pandas as pd
import geopandas as gpd
from openpyxl import load_workbook
from pathlib import Path
import plotly.graph_objects as go


# Load Mapbox token
mapbox_token = px.set_mapbox_access_token(open(".mapbox_token").read())


dataset_folder = Path('Datasets/')
workbook = load_workbook(dataset_folder / 'InteractiveMap_Data/InteractiveMap_Profile.xlsx')


# Map
p_choro = pd.read_excel(dataset_folder / 'InteractiveMap_Data/p_choro.xlsx')
p_choro['adm2_en'] = p_choro['adm2_en'].str.title()
p_choro['adm1_psgc'] = p_choro['adm1_psgc'].astype('int64')
p_choro['adm2_psgc'] = p_choro['adm2_psgc'].astype('int64')
p_choro['adm2_en'] = p_choro['adm2_en'].astype('object')
p_choro['geo_level'] = p_choro['geo_level'].astype('object')
p_choro['len_crs'] = p_choro['len_crs'].astype('int64')
p_choro['area_crs'] = p_choro['area_crs'].astype('int64')
p_choro['len_km'] = p_choro['len_km'].astype('int64')
p_choro['area_km2'] = p_choro['area_km2'].astype('float64')
p_choro['PROVINCE / LGU'] = p_choro['PROVINCE / LGU'].astype('object')
p_choro['2014'] = p_choro['2014'].astype('object')
p_choro['2015'] = p_choro['2015'].astype('object')
p_choro['2016'] = p_choro['2016'].astype('object')
p_choro['2017'] = p_choro['2017'].astype('object')
p_choro['2018'] = p_choro['2018'].astype('object')
p_choro['2019'] = p_choro['2019'].astype('object')
p_choro['2020'] = p_choro['2020'].astype('object')
p_choro['2021'] = p_choro['2021'].astype('object')
p_choro['2022'] = p_choro['2022'].astype('object')
p_choro['2023'] = p_choro['2023'].astype('object')
p_choro['_merge'] = p_choro['_merge'].astype('category')

province_options = [{'label': province, 'value': province} for province in p_choro['adm2_en'] if province is not None]


# LGU Profile
lgu_sheet = workbook['LGU']


lgu = []
category = []
percentage = []
province = []
revenue = []


for row in lgu_sheet.iter_rows(min_row=2, values_only=True):
   lgu.append(row[0])
   category.append(row[1])
   percentage.append(row[2])
   province.append(row[3])
   revenue.append(row[4])


lgu_data = [row[0].value for row in lgu_sheet.iter_rows(min_row=2, max_col=1)]
category_data = [row[1].value for row in lgu_sheet.iter_rows(min_row=2, max_col=2)]
province_data = [row[3].value for row in lgu_sheet.iter_rows(min_row=2, max_col=4)]
revenue_data = [row[4].value for row in lgu_sheet.iter_rows(min_row=2, max_col=5)]
lgu_options = [{'label': lgu_name, 'value': lgu_name} for lgu_name in lgu_data if lgu_name is not None]


def get_pillar_description(selected_pillar):
   pillar_descriptions = {
       'Resiliency': 'Applies to the capacity of a locality to build systems that can absorb change and disturbance and being able to adapt to such changes. It spans frameworks that bind LGUs and their constituents to prepare for possible shocks and stresses; budgeting for disaster risk reduction; hazard/risk identification mechanisms; resilience-related infrastructure; and resilience-related mechanisms.',
       'Government Efficiency': 'Refers to the quality and reliability of government services and government support for effective and sustainable productive expansion. This factor looks at government as an institution that is generally not corrupt; able to protect and enforce contracts; apply moderate and reasonable taxation and is able to regulate proactively.',
       'Innovation': 'Refers to the ability of a locality to harness its creative potential to improve or sustain current levels of productivity. It hinges mainly on the development of creative capital which are human resources, research capabilities, and networking capacities..',
       'Economic Dynamism': 'Refers to stable expansion of businesses and industries and higher employment. Matches output and productivity of the local economy with the local resources. Localities are centers of economic activities, and due to this, business expansion and job creation are easily observable in local settings.',
       'Infrastructure': 'Pertains to the physical assets that connect, expand, and sustain a locality and its surroundings to enable provision of goods and services. It involves basic inputs of production such as energy, water; interconnection of production such as transportation, roads and communications; sustenance of production such as waste, disaster preparedness, environmental sustainability; and human capital formation infrastructure.'
   }
   return pillar_descriptions.get(selected_pillar, 'No description available')


def get_lgu_province(selected_lgu):
   try:
       index = lgu_data.index(selected_lgu)
       return province_data[index]
   except ValueError:
       return 'No data available'


def get_lgu_category(selected_lgu):
   try:
       index = lgu_data.index(selected_lgu)
       return category_data[index]
   except ValueError:
       return 'No data available'


def get_lgu_revenue(selected_lgu):
   try:
       index = lgu_data.index(selected_lgu)
       return revenue_data[index]
   except ValueError:
       return 'No data available'


app = Dash(__name__, external_stylesheets=[dbc.themes.BOOTSTRAP])


app.layout = dbc.Container([
  
   dbc.Row([


   # First column
   dbc.Col([
       html.Div([
           html.H3('Interactive Map', style={'text-align': 'center', 'margin-bottom': '10px'}),
           dcc.Graph(id='choropleth-map')
       ])
   ]),
  
   # Second column
   dbc.Col([
       html.Div([
           html.H3('Province Profile', style={'text-align': 'center'}),
           html.Label('Select Province'),
           dcc.Dropdown(
               id='province-dropdown',
               options=province_options,
               value=[]
           ),
           html.H3('LGU Profile', style={'text-align': 'center'}),


           dbc.Row([


           # LGU Dropdown
           html.Label('Select LGU'),
           dcc.Dropdown(
               id='lgu-dropdown',
               options=lgu_options,
               value=[]
           ),
           dbc.Row([
               dbc.Col([
                   html.Label('Select Pillar'),
                   dcc.Dropdown(
                       id='pillar-dropdown',
                       options=[
                           {'label': 'Resiliency', 'value': 'Resiliency'},
                           {'label': 'Government Efficiency', 'value': 'Government Efficiency'},
                           {'label': 'Innovation', 'value': 'Innovation'},
                           {'label': 'Economic Dynamism', 'value': 'Economic Dynamism'},
                           {'label': 'Infrastructure', 'value': 'Infrastructure'},
                       ],
                       value=[]
                   ),
               ]),
               dbc.Col([
                   html.Label('Pillar Description'),
                   html.Div(id='pillar-description')
               ])
           ]),
           dbc.Row([
               dbc.Col([
                   html.Label('Province'),
                   html.Div(id='province-label')
               ]),
               dbc.Col([
                   html.Label('Category'),
                   html.Div(id='category-label')
               ]),
               dbc.Col([
                   html.Label('Revenue'),
                   html.Div(id='revenue-label')
               ])
           ])
       ])
   ]),
   ]),
  
   # Third column
   dbc.Col([
       html.Div([
           html.H3('Score per Pillar', style={'text-align': 'center', 'margin-bottom': '10px'}),
           dcc.Graph(id='bar-chart')
       ])
   ])
   ])
])


# Map
@app.callback(
   Output('choropleth-map', 'figure'),
   Input('province-dropdown', 'value')
)
def update_choropleth(selected_province):
   if not selected_province:
       raise dash.exceptions.PreventUpdate


   filtered_data = p_choro[p_choro['adm2_en'] == selected_province]


   if not filtered_data.empty:
       manila_center = {'lat': 14.5995, 'lon': 120.9842}


       fig = px.choropleth_mapbox(
           filtered_data,
           geojson=filtered_data.geometry,
           locations=filtered_data.index,
           color='2023', 
           hover_name='adm2_en',
           center=manila_center,
           zoom=4
       )


       fig.add_trace(go.Scattermapbox(
           lat=[filtered_data.geometry.centroid.y.iloc[0]],
           lon=[filtered_data.geometry.centroid.x.iloc[0]],
           mode='markers',
           marker=go.scattermapbox.Marker(
               size=10,
               color='red'
           ),
           hoverinfo='text',
           hovertext=f"Selected Province: {selected_province}"
       ))


       return fig
   else:
       return px.choropleth_mapbox()


# Descriptions
@app.callback(
   [
       Output('pillar-description', 'children'),
       Output('province-label', 'children'),
       Output('category-label', 'children'),
       Output('revenue-label', 'children'),
   ],
   [
       Input('lgu-dropdown', 'value'),
       Input('pillar-dropdown', 'value')
   ]
)
def update_labels(selected_lgu, selected_pillar):
   if selected_lgu and selected_pillar:
       pillar_description = get_pillar_description(selected_pillar)
       lgu_province = get_lgu_province(selected_lgu)
       lgu_category = get_lgu_category(selected_lgu)
       lgu_revenue = get_lgu_revenue(selected_lgu)


       return pillar_description, lgu_province, lgu_category, lgu_revenue
   else:
       return '', '', '', ''


# Bar Chart
@app.callback(
   Output('bar-chart', 'figure'),
   Input('lgu-dropdown', 'value')
)
def update_bar_chart(selected_lgu):
   if not selected_lgu:
       return {}


   lgu_index = lgu_data.index(selected_lgu) + 2
   lgu_data_row = list(lgu_sheet.iter_rows(min_row=lgu_index, max_row=lgu_index, min_col=6, max_col=10, values_only=True))[0]
   pillars = ['Resiliency', 'Government Efficiency', 'Innovation', 'Economic Dynamism', 'Infrastructure']


   fig = px.bar(
       x=pillars,
       y=lgu_data_row,
       labels={'x': 'Pillar', 'y': 'Score'},
       title=f'Scores per Pillar for {selected_lgu}'
   )


   return fig


if __name__ == '__main__':
   app.run_server(debug=True)