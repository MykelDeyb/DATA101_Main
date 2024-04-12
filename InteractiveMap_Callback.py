import dash
from dash import Dash, html, dcc, callback, Output, Input
import plotly.express as px
import dash_bootstrap_components as dbc
import pandas as pd
import geopandas as gpd
import numpy as np
from openpyxl import load_workbook
from pathlib import Path
import plotly.graph_objects as go
import numpy as np

dataset_folder = Path('Datasets/')
workbook = load_workbook(dataset_folder / 'InteractiveMap_Data/InteractiveMap_Profile.xlsx')

all_years = list(range(2014, 2024))

# Map
file_paths = [
    dataset_folder / "InteractiveMap_Data/Province JSON/provinces-region-bicolregionregionv.json",
    dataset_folder / "InteractiveMap_Data/Province JSON/provinces-region-autonomousregionofmuslimmindanaoarmm.json",
    dataset_folder / "InteractiveMap_Data/Province JSON/provinces-region-cagayanvalleyregionii.json",
    dataset_folder / "InteractiveMap_Data/Province JSON/provinces-region-calabarzonregioniva.json",
    dataset_folder / "InteractiveMap_Data/Province JSON/provinces-region-caragaregionxiii.json",
    dataset_folder / "InteractiveMap_Data/Province JSON/provinces-region-centralluzonregioniii.json",
    dataset_folder / "InteractiveMap_Data/Province JSON/provinces-region-centralvisayasregionvii.json",
    dataset_folder / "InteractiveMap_Data/Province JSON/provinces-region-cordilleraadministrativeregioncar.json",
    dataset_folder / "InteractiveMap_Data/Province JSON/provinces-region-davaoregionregionxi.json",
    dataset_folder / "InteractiveMap_Data/Province JSON/provinces-region-easternvisayasregionviii.json",
    dataset_folder / "InteractiveMap_Data/Province JSON/provinces-region-ilocosregionregioni.json",
    dataset_folder / "InteractiveMap_Data/Province JSON/provinces-region-metropolitanmanila.json",
    dataset_folder / "InteractiveMap_Data/Province JSON/provinces-region-mimaroparegionivb.json",
    dataset_folder / "InteractiveMap_Data/Province JSON/provinces-region-northernmindanaoregionx.json",
    dataset_folder / "InteractiveMap_Data/Province JSON/provinces-region-soccsksargenregionxii.json",
    dataset_folder / "InteractiveMap_Data/Province JSON/provinces-region-westernvisayasregionvi.json",
    dataset_folder / "InteractiveMap_Data/Province JSON/provinces-region-zamboangapeninsularegionix.json"
]


# Read each file and append to a list
ph_list = [gpd.read_file(file) for file in file_paths]

# Combine all GeoDataFrames in the list into a single GeoDataFrame
ph = gpd.GeoDataFrame(pd.concat(ph_list, ignore_index=True), crs=ph_list[0].crs)
#ph = ph.to_crs(epsg=32651)
p_score = pd.read_csv(dataset_folder / 'Province_Data/Overall Score.csv', encoding='latin1')
ph.loc[ph['PROVINCE'] == 'Agusan del Norte', 'PROVINCE'] = 'Agusan Del Norte'
ph.loc[ph['PROVINCE'] == 'Agusan del Sur', 'PROVINCE'] = 'Agusan Del Sur'
ph.loc[ph['PROVINCE'] == 'Batangas', 'PROVINCE'] = 'Batangas Province'
ph.loc[ph['PROVINCE'] == 'Biliran', 'PROVINCE'] = 'Biliran Province'
ph.loc[ph['PROVINCE'] == 'Cavite', 'PROVINCE'] = 'Cavite Province'
ph.loc[ph['PROVINCE'] == 'Cebu', 'PROVINCE'] = 'Cebu Province'
ph.loc[ph['PROVINCE'] == 'North Cotabato', 'PROVINCE'] = 'Cotabato (North Cotabato)'
ph.loc[ph['PROVINCE'] == 'Davao de Oro', 'PROVINCE'] = 'Davao De Oro'
ph.loc[ph['PROVINCE'] == 'Davao del Norte', 'PROVINCE'] = 'Davao Del Norte'
ph.loc[ph['PROVINCE'] == 'Davao del Sur', 'PROVINCE'] = 'Davao Del Sur'
ph.loc[ph['PROVINCE'] == 'Iloilo', 'PROVINCE'] = 'Iloilo Province'
ph.loc[ph['PROVINCE'] == 'Lanao del Norte', 'PROVINCE'] = 'Lanao Del Norte'
ph.loc[ph['PROVINCE'] == 'Lanao del Sur', 'PROVINCE'] = 'Lanao Del Sur'
ph.loc[ph['PROVINCE'] == 'Leyte', 'PROVINCE'] = 'Leyte Province'
ph.loc[ph['PROVINCE'] == 'Masbate', 'PROVINCE'] = 'Masbate Province'
ph.loc[ph['PROVINCE'] == 'Romblon', 'PROVINCE'] = 'Romblon Province'
ph.loc[ph['PROVINCE'] == 'Samar', 'PROVINCE'] = 'Samar (Western Samar)'
ph.loc[ph['PROVINCE'] == 'Siquijor', 'PROVINCE'] = 'Siquijor Province'
ph.loc[ph['PROVINCE'] == 'Sorsogon', 'PROVINCE'] = 'Sorsogon Province'
ph.loc[ph['PROVINCE'] == 'Surigao del Norte', 'PROVINCE'] = 'Surigao Del Norte'
ph.loc[ph['PROVINCE'] == 'Surigao del Sur', 'PROVINCE'] = 'Surigao Del Sur'
ph.loc[ph['PROVINCE'] == 'Tarlac', 'PROVINCE'] = 'Tarlac Province'
ph.loc[ph['PROVINCE'] == 'Zamboanga del Norte', 'PROVINCE'] = 'Zamboanga Del Norte'
ph.loc[ph['PROVINCE'] == 'Zamboanga del Sur', 'PROVINCE'] = 'Zamboanga Del Sur'
#ph.loc[ph['PROVINCE'] == 'Maguindanao del Norte', 'PROVINCE'] = 'Maguindanao'
#ph.loc[ph['PROVINCE'] == 'Maguindanao del Sur', 'PROVINCE'] = 'Maguindanao'
ph.loc[ph['PROVINCE'] == 'Metropolitan Manila', 'PROVINCE'] = 'Metro Manila'
p_choro = pd.merge(ph, p_score,  left_on='PROVINCE', right_on='PROVINCE / LGU', how='left', indicator=True)

province_options = [{'label': province, 'value': province} for province in p_choro['PROVINCE'] if province is not None]

# Province Profile 

province_sheet = workbook['Province']

province = []
region = []
population = []
province_revenue = []
rank = []

for row in province_sheet.iter_rows(min_row=2, values_only=True):
   region.append(row[1])
   population.append(row[2])
   province_revenue.append(row[3])
   rank.append(row[4])

province_data = [row[0].value for row in province_sheet.iter_rows(min_row=2)]
region_data = [row[1].value for row in province_sheet.iter_rows(min_row=2)]
population_data = [row[2].value for row in province_sheet.iter_rows(min_row=2)]
province_revenue_data = [row[3].value for row in province_sheet.iter_rows(min_row=2)]
rank_data = [row[4].value for row in province_sheet.iter_rows(min_row=2)]

def get_province_region(province):
   try:
       index = province_data.index(province)
       return region_data[index]
   except ValueError:
       return 'No data available'
   
def get_province_population(province):
   try:
       index = province_data.index(province)
       return population_data[index]
   except ValueError:
       return 'No data available'


def get_province_revenue(province):
   try:
       index = province_data.index(province)
       return province_revenue[index]
   except ValueError:
       return 'No data available'


def get_province_rank(province):
   try:
       index = province_data.index(province)
       return rank_data[index]
   except ValueError:
       return 'No data available'
   
# LGU Profile
lgu_sheet = workbook['LGU']

lgu = []
category = []
percentage = []
lgu_province = []
revenue = []

for row in lgu_sheet.iter_rows(min_row=2, values_only=True):
   lgu.append(row[0])
   category.append(row[1])
   percentage.append(row[2])
   lgu_province.append(row[3])
   revenue.append(row[4])

lgu_data = [row[0].value for row in lgu_sheet.iter_rows(min_row=2, max_col=1)]
category_data = [row[1].value for row in lgu_sheet.iter_rows(min_row=2, max_col=2)]
lgu_province_data = [row[3].value for row in lgu_sheet.iter_rows(min_row=2, max_col=4)]
revenue_data = [row[4].value for row in lgu_sheet.iter_rows(min_row=2, max_col=5)]
lgu_options = [{'label': lgu_name, 'value': lgu_name} for lgu_name in lgu_data if lgu_name is not None]

def get_lgu_province(selected_lgu):
   try:
       index = lgu_data.index(selected_lgu)
       return lgu_province_data[index]
   except ValueError:
       return 'No data available'

def get_lgu_category(selected_lgu):
   try:
       index = lgu_data.index(selected_lgu)
       return category_data[index]
   except ValueError:
       return 'No data available'

def get_lgu_revenue(selected_lgu):
   try:
       index = lgu_data.index(selected_lgu)
       return revenue_data[index]
   except ValueError:
       return 'No data available'


app = Dash(__name__, external_stylesheets=[dbc.themes.BOOTSTRAP])

app.layout = dbc.Container([
  
   dbc.Row([
       
   # First column
   dbc.Col([
       # Card 1, Entire First Column
       dbc.Card([
           dbc.CardBody([
           dbc.Row([
                dbc.Row([
                    html.Div([
                    html.H3('Choropleth Map', style={'text-align': 'center', 'margin-bottom': '10px'}),
                    dbc.Row([
                        html.Label('Select Year', style={'font-weight': 'bold'}),
                        dcc.Dropdown(
                            id='map-year-dropdown-province',
                            options=[{'label': str(year), 'value': year} for year in all_years],
                            value=2023,
                            style={'width': '80px', 'margin-bottom': '1px', 'display': 'inline-block', 'vertical-align': 'middle'}
                        )])
                        
                    
                    
                ]),
                    dcc.Graph(id='choropleth-map')  # Display the initial choropleth map
                ])
            ]),
            ], style={'height': '953px'})
       ], color='light')
       
       
   ]),
  
   # Second column
   dbc.Col([
       html.Div([
           # Card 2, Top Right
           dbc.Card([dbc.CardBody([
                dbc.Row([
                html.H3('Province Profile', style={'text-align': 'center'}),
                html.Label('Select Province', style={'font-weight': 'bold'}),
                dcc.Dropdown(
                    id='province-dropdown',
                    options=sorted(province_options, key=lambda d: d['label']),
                    value=[]
                ),
                dbc.Row([
                    dbc.Col([
                        html.Div(id='map_prov_table')
                    ])
                    
                ], justify='center',align='center', style={'margin-bottom': '10px', 'margin-top':'10px'})
                ])
           ], style={'height': '220px'})], color='light'),

           # Card 3, Bottom Right
           dbc.Card([dbc.CardBody([
               dbc.Row([
               html.H3('LGU Profile', style={'text-align': 'center'}),
                # LGU Dropdown
                html.Label('Select LGU', style={'font-weight': 'bold'}),
                dcc.Dropdown(
                    id='lgu-dropdown',
                    options=lgu_options,
                    value=[]
                ),
                dbc.Row([
                    dbc.Col([
                        html.Div(id='map_lgu_table')
                    ])
                ], justify='center',align='center', style={'margin-bottom': '10px', 'margin-top':'10px'}),
                dbc.Row([
                    html.Div([
                        dcc.Graph(id='bar-chart-map')
                    ])
                ]),
                dbc.Row([
                    dbc.Col([
                        html.Label('Select Pillar', style={'font-weight': 'bold'}),
                        dcc.Dropdown(
                            id='pillar-dropdown-map',
                            options=[
                                {'label': 'Resiliency', 'value': 'Resiliency'},
                                {'label': 'Government Efficiency', 'value': 'Government Efficiency'},
                                {'label': 'Innovation', 'value': 'Innovation'},
                                {'label': 'Economic Dynamism', 'value': 'Economic Dynamism'},
                                {'label': 'Infrastructure', 'value': 'Infrastructure'},
                            ],
                            value=[]
                        ),
                    ]),
                    dbc.Col([
                        html.Label('Pillar Description', style={'font-weight': 'bold'}),
                        html.Div(id='pillar-description')
                    ])
                ]),
            ])
           ], style={'height': '720px'})], color='light', style={'margin-top': '10px'})
           
   ]),
   ])
   ])
], style={'margin-top': '20px'})

# Map Page, Table Province 
@app.callback(
    [
        Output('map_prov_table','children')
    ],
    [
        Input('province-dropdown','value')
    ]
)
def update_labels(province):
    if province:
       province_region = get_province_region(province)
       province_population = get_province_population(province)
       province_revenue = get_province_revenue(province)
       province_rank = get_province_rank(province)

       table_rows = [
                    html.Tr([
                            html.Th('Region', style={'border-bottom': '1px solid #ddd', 'font-size': '14px', 'text-align': 'center', 'background-color': '#F1F1F1', 'width': '25%'}),  
                            html.Th('Population', style={'border-bottom': '1px solid #ddd', 'font-size': '14px', 'text-align': 'center', 'background-color': '#F1F1F1', 'width': '25%'}),
                            html.Th('Revenue', style={'border-bottom': '1px solid #ddd', 'font-size': '14px', 'text-align': 'center', 'background-color': '#F1F1F1', 'width': '25%'}),
                            html.Th('Ranking', style={'border-bottom': '1px solid #ddd', 'font-size': '14px', 'text-align': 'center', 'background-color': '#F1F1F1', 'width': '25%'})
                            ], style={'width': '100%'})
                    ]
       table_rows.append(html.Tr([
            html.Td(province_region, style={'border-bottom': '1px solid #ddd', 'font-size': '14px', 'text-align': 'center', 'width': '25%'}),  
            html.Td(province_population, style={'border-bottom': '1px solid #ddd', 'font-size': '14px', 'text-align': 'center', 'width': '25%'}),
            html.Td(province_revenue, style={'border-bottom': '1px solid #ddd', 'font-size': '14px', 'text-align': 'center', 'width': '25%'}),
            html.Td(province_rank, style={'border-bottom': '1px solid #ddd', 'font-size': '14px', 'text-align': 'center', 'width': '25%'})
        ], style={'width': '100%'}))
       return [html.Table(table_rows, style={'width': '100%', 'margin': 'auto', 'textAlign': 'center'})]  # Center the table within the card
    else:
        return [[]]  # Return an empty list to match the expected output type
# Map LGU Table
@app.callback(
    [
        Output('map_lgu_table','children')
    ],
    [
        Input('lgu-dropdown','value')
    ]
)
def update_labels(selected_lgu):
    if selected_lgu:
       lgu_province = get_lgu_province(selected_lgu)
       lgu_category = get_lgu_category(selected_lgu)
       lgu_revenue = get_lgu_revenue(selected_lgu)

       table_rows = [
                    html.Tr([
                            html.Th('Province', style={'border-bottom': '1px solid #ddd', 'font-size': '14px', 'text-align': 'center', 'background-color': '#F1F1F1', 'width': '34%'}),  
                            html.Th('Category', style={'border-bottom': '1px solid #ddd', 'font-size': '14px', 'text-align': 'center', 'background-color': '#F1F1F1', 'width': '33%'}),
                            html.Th('Revenue', style={'border-bottom': '1px solid #ddd', 'font-size': '14px', 'text-align': 'center', 'background-color': '#F1F1F1', 'width': '33%'})
                            ], style={'width': '100%'})
                    ]
       table_rows.append(html.Tr([
            html.Td(lgu_province, style={'border-bottom': '1px solid #ddd', 'font-size': '14px', 'text-align': 'center', 'width': '34%'}),  
            html.Td(lgu_category, style={'border-bottom': '1px solid #ddd', 'font-size': '14px', 'text-align': 'center', 'width': '33%'}),
            html.Td(lgu_revenue, style={'border-bottom': '1px solid #ddd', 'font-size': '14px', 'text-align': 'center', 'width': '33%'})
        ], style={'width': '100%'}))
       return [html.Table(table_rows, style={'width': '100%', 'margin': 'auto', 'textAlign': 'center'})]  # Center the table within the card
    else:
        return [[]]  # Return an empty list to match the expected output type

def get_pillar_description(selected_pillar):
   pillar_descriptions = {
       'Resiliency': 'Applies to the capacity of a locality to build systems that can absorb change and disturbance and being able to adapt to such changes',
       'Government Efficiency': 'Refers to the quality and reliability of government services and government support for effective and sustainable productive expansion',
       'Innovation': 'Refers to the ability of a locality to harness its creative potential to improve or sustain current levels of productivity',
       'Economic Dynamism': 'Refers to stable expansion of businesses and industries and higher employment',
       'Infrastructure': 'Pertains to the physical assets that connect, expand, and sustain a locality and its surroundings to enable provision of goods and services'
   }
   return pillar_descriptions.get(selected_pillar, 'No description available')

@app.callback(
    Output('pillar-description', 'children'),
    [Input('pillar-dropdown-map', 'value')]
)
def update_pillar_description(selected_pillar):
    if not selected_pillar:
        return "-" 
    else:
        return get_pillar_description(selected_pillar)

# Bar Chart
@app.callback(
   Output('bar-chart-map', 'figure'),
   Input('lgu-dropdown', 'value')
)
def update_bar_chart(selected_lgu):
   if not selected_lgu:
       return {}

   lgu_index = lgu_data.index(selected_lgu) + 2
   lgu_data_row = list(lgu_sheet.iter_rows(min_row=lgu_index, max_row=lgu_index, min_col=6, max_col=10, values_only=True))[0]
   pillars = ['Resiliency', 'Government Efficiency', 'Innovation', 'Economic Dynamism', 'Infrastructure']

   fig = px.bar(
       y=lgu_data_row,
       labels={'x': 'Pillar', 'y': 'Score'},
       color=pillars,
       height=400,
       title=f'Scores per Pillar for {selected_lgu}'
   )
   fig.update_layout(
            showlegend=True,
            title_font=dict(size=20, family='Arial Black'),
        )
    
   fig.update_traces(hovertemplate='Pillar: %{x} <br>Score: $%{y}<extra></extra>')
   return fig

# Map
@app.callback(
    Output('choropleth-map', 'figure'),
    Input('map-year-dropdown-province', 'value'),
    Input('province-dropdown', 'value')
)
def update_choropleth(map_year, province):
    initial_column_values = p_choro.set_index('PROVINCE')[str(map_year)].replace('-', np.nan).astype(float).fillna(0)

    initial_fig = px.choropleth_mapbox(
        p_choro,
        geojson=p_choro,
        locations=p_choro.PROVINCE,
        featureidkey="properties.PROVINCE",
        color=initial_column_values,
        color_continuous_scale='Viridis',
        hover_name='PROVINCE',  
        hover_data={str(map_year): True},
        labels={map_year: 'Overall CMCI Score'},
        center={'lat': 12.8797, 'lon': 121.7740},  # Center coordinates of the Philippines
        mapbox_style="carto-positron",  # Choose the Mapbox map style
        zoom=5  # Adjust the initial zoom level as needed
    )

    initial_fig.update_layout(
        title='Choropleth Map',
        margin={"r": 0, "t": 0, "l": 0, "b": 0},
        height=800,
    )
    
    initial_fig.update_traces(hovertemplate='<b>%{hovertext}</b><br>CMCI Score: %{customdata}'
    
    )

    lon_manila = ph.loc[ph['PROVINCE'] == "Metro Manila", 'geometry'].get_coordinates().iloc[0]['x']
    lat_manila = ph.loc[ph['PROVINCE'] == "Metro Manila", 'geometry'].get_coordinates().iloc[0]['y']

    initial_fig.add_scattermapbox(
                lat=[lat_manila],
                lon=[lon_manila],
                mode='markers',
                text="Coordinates",
                marker_size=15,
                opacity=0.8,
                marker_color='rgb(235, 0, 100)',
                showlegend=False,
                name=""
            )

    if province: 
        lon_selected = ph.loc[ph['PROVINCE'] == province, 'geometry'].get_coordinates().iloc[0]['x']
        lat_selected = ph.loc[ph['PROVINCE'] == province, 'geometry'].get_coordinates().iloc[0]['y']

        initial_fig.add_scattermapbox(
            lat=[lat_manila],
            lon=[lon_manila],
            mode='markers',
            text="Manila",
            marker_size=15,
            opacity=0.8,
            marker_color='rgb(235, 0, 100)',
            showlegend=False,
            name=""
        )

        initial_fig.add_scattermapbox(
            lat=[lat_manila, lat_selected],
            lon=[lon_manila, lon_selected],
            mode='markers+lines',  # Include lines in the mode
            text=["Manila", province],  # Adjust the text accordingly
            marker_size=15,
            opacity=0.8,
            marker_color='rgb(235, 0, 100)',
            showlegend=False,
            name=""
        )

    initial_fig.add_trace(initial_fig.data[0])

    return initial_fig


if __name__ == '__main__':
   app.run_server(debug=True)

